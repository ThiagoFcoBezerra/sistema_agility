import datetime
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook
from orcamento.models import Categoria, Lancamento, Orcamento, Faturamento, Autorizacao
from django.db.models.aggregates import Sum
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import permission_classes, api_view
from rest_framework.response import Response
from rest_framework import viewsets
from orcamento.serializers import CategoriaSerializer

@login_required
def index(request):
    
    return render(request, 'orcamento/index.html', contexto_index(request))

@api_view(['GET'])
def api_index(request):
    dados = contexto_index(request)
    contexto = []
    for k in dados:
        if k == 'orcamento_total_mensal' or k == 'lancamento_total_mensal' or k == 'orcamento_total_anual' or k == 'lancamento_total_anual' or k == 'autorizacao_total_mensal':
            obj = {}
            lista = []
            for dado in dados[k]:
                lista.append(dado)
            obj[k] = lista
            contexto.append(obj)
        if k == 'mes' or k == 'data_selecionada':
            obj = {}
            obj[k] = dados[k]
            contexto.append(obj)
        if k == 'categorias':
            lista = []
            obj = {}
            for categoria in dados[k]:
                obj_cat = {'categoria': categoria.nome}
                lista.append(obj_cat)
            obj[k] = lista
            contexto.append(obj)
                
            
    return Response(contexto) 


def contexto_index(request):
    if request.method == 'POST':
            data_selecionada = request.POST['data']
            ano = data_selecionada[0:4]
            mes = data_selecionada[5:]
    else:
        data = datetime.datetime.now()
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
        data_selecionada = f'{ano}-{mes}'


    orcamento_total_mensal = Orcamento.objects.filter(data_orcamento__month = mes, data_orcamento__year = ano).values('categoria__nome').annotate(total_mes = Sum('valor_orc'))
    lancamento_total_mensal = Lancamento.objects.filter(data_pagamento__month = mes, data_pagamento__year = ano).values('categoria__nome').annotate(total_mes = Sum('valor_pago'))
    orcamento_total_anual = Orcamento.objects.filter(data_orcamento__year = ano).values('categoria__nome').annotate(total_mes = Sum('valor_orc'))
    lancamento_total_anual = Lancamento.objects.filter(data_pagamento__year = ano).values('categoria__nome').annotate(total_mes = Sum('valor_pago'))
    autorizacao_total_mensal = Autorizacao.objects.filter(
        data_autorizacao__month = mes, 
        data_autorizacao__year = ano, 
        utilizada = False,
        autorizacao_status = 'APR'
        ).values('categoria__nome').annotate(total_mes = Sum('valor_autorizacao'))
    categorias = Categoria.objects.all().order_by('nome')

    context = {
        'orcamento_total_mensal' : orcamento_total_mensal,
        'lancamento_total_mensal' : lancamento_total_mensal,
        'orcamento_total_anual' : orcamento_total_anual,
        'lancamento_total_anual' : lancamento_total_anual,
        'autorizacao_total_mensal' : autorizacao_total_mensal,
        'categorias' : categorias,
        'data_selecionada' : data_selecionada,
        'mes': mes
    }
    return context

@login_required
def testa_arquivo(request):
    
    if request.method == 'POST':
        arquivo = request.FILES['arquivo']
        planilha = request.POST['planilha']
        col_data = request.POST['col_data'].upper()
        col_descricao = request.POST['col_descricao'].upper()
        col_valor = request.POST['col_valor'].upper()
        col_categoria = request.POST['col_categoria'].upper()

        wb = load_workbook(arquivo)

        ws = wb.get_sheet_by_name(planilha)
        mes = ''
        for data in ws[col_data]:
            if type(data.value) == datetime.datetime:
                if data.value.strftime('%m') != mes:
                    lancamentos_a_excluir = Lancamento.objects.filter(data_pagamento__month = data.value.strftime("%m"), data_pagamento__year = data.value.strftime("%Y"))
                    for lancamento in lancamentos_a_excluir:
                        print(lancamento)
                        lancamento.delete()
                        mes = data.value.strftime('%m')

        
        for cell1, cell2, cell3, cell4 in zip(ws[col_categoria], ws[col_descricao], ws[col_valor], ws[col_data]):
            if type(cell3.value) == type(3.14) and cell1.value != None and cell2.value != None and cell4.value != None:
                
                cat = Categoria.objects.filter(nome = cell1.value)
                
                if not cat:
                    categoria = Categoria.objects.create(nome = cell1.value)
                else:
                    categoria = cat[0]
                valor_pago = round(cell3.value,2) 
                lanca = Lancamento.objects.create(categoria = categoria, descricao = cell2.value, valor_pago = valor_pago, data_pagamento = cell4.value)
                lanca.save()
            
    return render(request, 'orcamento/atualiza_lancamentos2.html')

@login_required
def orcamento(request):
    
    data_atual = datetime.datetime.now()
    
    if request.method == 'POST':
        data_buscada = request.POST['mes_buscado']
        mes_buscado = data_buscada[5:7]
        ano_buscado = data_buscada[0:4]
    else:
        mes_buscado = data_atual.strftime('%m')
        ano_buscado = data_atual.strftime('%Y')

    data_buscada = f'{ano_buscado}-{mes_buscado}'
    categorias = Categoria.objects.all().order_by('nome')
    orc = Orcamento.objects.filter(data_orcamento__month = mes_buscado, data_orcamento__year = ano_buscado)

    for cat in categorias:
        orcamento_existe = orc.filter(categoria = cat)
        if not orcamento_existe:
            Orcamento.objects.create(categoria = cat, valor_orc = 0, data_orcamento = f'{ano_buscado}-{mes_buscado}-01')

    orcamentos = Orcamento.objects.filter(data_orcamento__month = mes_buscado, data_orcamento__year = ano_buscado).values('categoria__nome').annotate(soma = Sum('valor_orc')).order_by('categoria__nome')
    
    for orcamento in orcamentos:
        orcamento['soma'] = float(orcamento['soma'])
        orcamento['soma'] = round(orcamento['soma'], 2)

    dados = {
        'orcamentos': orcamentos,
        'data': data_buscada
    }
    return render(request, 'orcamento/orcamento2.html', dados)
    
@login_required
@permission_required('orcamento.change_orcamento', raise_exception=True)
def cadastra_orcamento(request):
    data = request.POST['data']
    mes = int(data[5:7])
    ano = int(data[0:4])
    if request.POST.get('recorrencia'):
        mes_final = 13
    else:
        mes_final = mes+1

    for m in range(mes, mes_final):
        m = str(m)
        if len(m) < 2:
            m = '0'+m

        data = f'{ano}-{m}-01'
        
        for cat, valor in zip(request.POST.getlist('categoria'), request.POST.getlist('valor')):
            print(valor)
            #valor = valor.replace('.','')
            #valor = valor.replace(',','.')
            #valor = float(valor)
            #valor = round(valor,2)
            
            orc = Orcamento.objects.filter(categoria = Categoria.objects.get(nome = cat), data_orcamento = data)
            if not orc:
                Orcamento.objects.create(categoria = Categoria.objects.get(nome = cat), data_orcamento = data, valor_orc = valor)
            else:
                #orc[0].categoria = Categoria.objects.get(nome = cat)
                #orc[0].data_orcamento = data
                orc[0].valor_orc = valor
                orc[0].save()


    return redirect('index')

@login_required
def detalhes(request, id, m):

    if request.method == 'POST':
        data_selecionada = request.POST['data']
        ano = data_selecionada[0:4]
        mes = data_selecionada[5:]
        cat_id = id
    else:
        data = datetime.datetime.now()
        ano = data.strftime("%Y")
        cat_id = id
        mes = m
        if m < 10:
            mes = f'0{m}'
        data_selecionada = f'{ano}-{mes}'

    categorias = Categoria.objects.all().order_by('nome')
    categoria = categorias.get(pk = cat_id)
    lancamentos_a_exibir = Lancamento.objects.filter(categoria__id = id, data_pagamento__month = mes, data_pagamento__year = ano)
    lancamento_total_mensal = Lancamento.objects.filter(categoria__id = id, data_pagamento__year = ano).values('data_pagamento__month').annotate(total=Sum('valor_pago')).order_by('data_pagamento__month')
    orcamento_total_mensal = Orcamento.objects.filter(categoria__id = id, data_orcamento__year = ano).values('data_orcamento__month').annotate(total=Sum('valor_orc')).order_by('data_orcamento__month')

    dados = {
        'mes': mes,
        'categorias':categorias,
        'data_selecionada':data_selecionada,
        'categoria': categoria,
        'id':cat_id,
        'lancamentos_a_exibir':lancamentos_a_exibir,
        'lancamento_total_mensal':lancamento_total_mensal,
        'orcamento_total_mensal':orcamento_total_mensal,
    }
    return render(request, 'orcamento/detalhes.html', dados)

@login_required
def cadastra_faturamento(request):
    if request.method == 'POST':
        data = request.POST['data']
        valor = request.POST['valor']

        Faturamento.objects.create(faturamento_data = data, faturamento_valor = valor)
    
    return redirect('resultado')

@login_required
def resultado(request):
    if request.method == 'POST':
        data_selecionada = request.POST['data_sel']
        ano = data_selecionada[0:4]
        mes = data_selecionada[5:]
    else:
        data = datetime.datetime.now()
        ano = data.strftime("%Y")
        mes = data.strftime("%m")
    data_selecionada = f'{ano}-{mes}'

    faturamentos = Faturamento.objects.filter(faturamento_data__month = mes, faturamento_data__year = ano).order_by('faturamento_data')
    fat_dia = Faturamento.objects.filter(faturamento_data__month = mes, faturamento_data__year = ano).values('faturamento_data__day').annotate(total=Sum('faturamento_valor')).order_by('faturamento_data__day')
    fat_mes = Faturamento.objects.filter(faturamento_data__year = ano).values('faturamento_data__month').annotate(total=Sum('faturamento_valor')).order_by('faturamento_data__month')
    pag_dia = Lancamento.objects.filter(data_pagamento__month = mes, data_pagamento__year = ano).values('data_pagamento__day').annotate(total=Sum('valor_pago')).order_by('data_pagamento__day')
    pag_mes = Lancamento.objects.filter(data_pagamento__year = ano).values('data_pagamento__month').annotate(total=Sum('valor_pago')).order_by('data_pagamento__month')

    dados = {
        'faturamentos':faturamentos,
        'fat_dia':fat_dia,
        'fat_mes':fat_mes,
        'pag_dia':pag_dia,
        'pag_mes':pag_mes,
        'data_selecionada':data_selecionada,
    }

    return render(request, 'orcamento/resultado.html', dados)

@login_required
def exclui_faturamento(request, id):
    faturamento = Faturamento.objects.get(pk = id)
    faturamento.delete()

    return redirect('resultado')

@login_required
def atualiza_faturamento(request, id):
    if request.method == 'POST':
        data = request.POST['data']
        valor = request.POST['valor']

        faturamento = Faturamento.objects.get(pk=id)
        faturamento.faturamento_data = data
        faturamento.faturamento_valor = valor

        faturamento.save() 
    return redirect('resultado')

@login_required
@permission_required('orcamento.add_autorizacao')
def cria_autorizacao(request):
    if request.method == 'POST':
        categoria_id = request.POST['id']
        justificativa = request.POST['justificativa']
        valor = request.POST['valor']
        user = user = get_object_or_404(User,pk=request.user.id)
        Autorizacao.objects.create(
            categoria = Categoria.objects.get(pk = categoria_id),
            justificativa = justificativa,
            valor_autorizacao = valor,
            solicitante = user)

    return redirect('lista_autorizacao')

@login_required
def carrega_autorizacao(request, id):
    autorizacao = Autorizacao.objects.get(pk = id)
    categorias = Categoria.objects.all().order_by('nome')

    context = {
        'autorizacao': autorizacao,
        'categorias': categorias
    }

    return render(request, 'orcamento/form_autorizacao.html', context)

@login_required
@permission_required('orcamento.change_autorizacao')
def edita_autorizacao(request):
    if request.method == 'POST':
        autorizacao = Autorizacao.objects.get(pk = request.POST['id'])
        autorizacao.categoria = get_object_or_404(Categoria, pk = request.POST['categoria'])
        autorizacao.justificativa = request.POST['justificativa']
        autorizacao.valor_autorizacao = request.POST['valor']
        autorizacao.solicitante = get_object_or_404(User,pk=request.user.id)
        autorizacao.save()

    return redirect('lista_autorizacao')

@login_required
@permission_required('orcamento.delete_autorizacao')
def exclui_autorizacao(request, id):

    autorizacao = Autorizacao.objects.get(pk=id)
    autorizacao.delete()

    return redirect('lista_autorizacao')

@login_required
@permission_required('orcamento.pode_despachar', raise_exception=True)
def despacha_autorizacao(request):

    if request.method == 'POST':

        autorizacao = Autorizacao.objects.get(pk=request.POST['id'])
        autorizacao.autorizacao_status = request.POST['status']
        autorizacao.despacho = request.POST['despacho']
        autorizacao.despachante = get_object_or_404(User, pk=request.user.id)
        autorizacao.save()
        return redirect('lista_autorizacao')

@login_required
def lista_autorizacao(request):
    data = request.GET.get("data")
    if not data:
        data = datetime.datetime.now()
    else:
        data = datetime.date(int(data[0:4]), int(data[5:7]), 1)
    autorizacoes = Autorizacao.objects.filter(data_autorizacao__month = data.strftime('%m'), data_autorizacao__year = data.strftime('%Y')).order_by('-data_criacao')
    aguardando = autorizacoes.filter(autorizacao_status = 'AGU')
    despachadas = autorizacoes.exclude(autorizacao_status = 'AGU')
    pag_aguardando = Paginator(aguardando, 10)
    pag_despachadas = Paginator(despachadas, 10)
    page_number_aguardando = request.GET.get("page_agu")
    page_number_despachada = request.GET.get("page_des")
    page_aguardando = pag_aguardando.get_page(page_number_aguardando)
    page_despachada = pag_despachadas.get_page(page_number_despachada)

    context = {
        'page_aguardando':page_aguardando,
        'page_despachada':page_despachada,
        'data': data,
    }
    return render (request, 'orcamento/lista_autorizacoes.html', context)

@login_required
@permission_required('orcamento.view_autorizacao')
def detalha_autorizacao(request, id):
    autorizacao = Autorizacao.objects.get(pk = id)

    context = {
        'autorizacao': autorizacao,
    }

    return render(request, 'orcamento/autorizacao_detalhes.html', context)

def utiliza_autorizacao(request):
    if request.method == 'POST':
        aut = Autorizacao.objects.get(pk = request.POST['id'])
        if aut.autorizacao_status == 'APR' and aut.utilizada == False:
            aut.descreve_utilizacao = request.POST['justificativa']
            aut.utilizada = True
            aut.save()


    return redirect('lista_autorizacao')

class CategoriaViewSet(viewsets.ModelViewSet):
    """ Retorna todas as categorias """
    
    queryset = Categoria.objects.all().order_by('nome')
    serializer_class = CategoriaSerializer
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getTotalLancamento(request, id, m, a):
    total_lancamento = Lancamento.objects.filter(categoria = id, data_pagamento__month = m, data_pagamento__year = a).values('categoria__nome').annotate(total = Sum('valor_pago'))
    return Response(total_lancamento)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getTotalOrcamento(request, id, m, a):
    total_orcamento = Orcamento.objects.filter(categoria = id, data_orcamento__month = m, data_orcamento__year = a).values('categoria__nome').annotate(total = Sum('valor_orc'))
    
    return Response(total_orcamento)